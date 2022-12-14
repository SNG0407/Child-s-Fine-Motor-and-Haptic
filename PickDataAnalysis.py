import numpy as np
#import pandas as pd
from dtaidistance import dtw
import matplotlib.pyplot as plt


import openpyxl
from random import *

wb = openpyxl.load_workbook('/Users/rlady/Desktop/소융캡디 데이터/DataList.xlsx')
print(wb.sheetnames)

EMG_Grab = wb['Pickpy']


#ColumnList = [0 for i in range(59)]
Num = [0 for i in range(250)]
LeapMotion = [0 for i in range(250)]
real = [0 for i in range(250)]
touch = [0 for i in range(250)]
glove = [0 for i in range(250)]
haptic = [0 for i in range(250)]

DataList =[Num, LeapMotion, real, touch, glove, haptic]
#print(DataList)
#print(EMG_Grab['A2'].value)

FirstCount = 0
secondCount = 0
#Column 세로 값 모두 읽기
for cols in EMG_Grab.iter_cols():

    for cell in cols:
        print(cell.value, end=' ')
        DataList[FirstCount][secondCount] = cell.value
        #print(secondCount, end=' ')
        secondCount = secondCount + 1
    #print(FirstCount)
    secondCount = 0
    FirstCount = FirstCount + 1
    print()

#print(DataList[1])

LeapMotion_G = DataList[1][1:60]# / 11.16949153
LeapMotion_R = DataList[1][61: 60+59]#/30.28813559
LeapMotion_B = DataList[1][60+61: 60+60+59]#/3.762711864
LeapMotion_Y = DataList[1][60+60+61: 60+60+60+59]#/8.898305085
print()
print("LeapMotion Data")
print(LeapMotion_G)
print(LeapMotion_R)
print(LeapMotion_B)
print(LeapMotion_Y)

real_G = DataList[2][1:60]#/13.55932203
real_R = DataList[2][61: 60+59]#/22.45762712
real_B = DataList[2][60+61: 60+60+59]#/9.915254237
real_Y = DataList[2][60+60+61: 60+60+60+59]#/11.74576271
print()
print("Real Data")
print(real_G)
print(real_R)
print(real_B)
print(real_Y)

touch_G = DataList[3][1:60]#/11.6440678
touch_R = DataList[3][61: 60+59]#/18.13559322
touch_B = DataList[3][60+61: 60+60+59]#/20.23728814
touch_Y = DataList[3][60+60+61: 60+60+60+59]#/28.16949153
print()
print("Touch Data")
print(touch_G)
print(touch_R)
print(touch_B)
print(touch_Y)

glove_G = DataList[4][1:60]#/12.88135593
glove_R = DataList[4][61: 60+59]#/10.33898305
glove_B = DataList[4][60+61: 60+60+59]#/13.10169492
glove_Y = DataList[4][60+60+61: 60+60+60+59]#/14.08474576
print()
print("Glove Data")
print(glove_G)
print(glove_R)
print(glove_B)
print(glove_Y)

haptic_G = DataList[5][1:60]#/16.74576271
haptic_R = DataList[5][61: 60+59]#/39.83050847
haptic_B = DataList[5][60+61: 60+60+59]#/17.49152542
haptic_Y = DataList[5][60+60+61: 60+60+60+59]#/59.76271186
print()
print("Haptic Data")
print(haptic_G)
print(haptic_R)
print(haptic_B)
print(haptic_Y)

#a = [1, 2, 5, 7, 4, 3, 6, 8, 2, 1]
#b = [3, 6, 1, 2, 8, 9, 3, 4, 3, 2, 1, 2]
#c = [2, 5, 7, 4, 3, 6, 8, 2, 1, 1]
print()
print("DTW Pick 분석")
distance_Grab_R_L_G = dtw.distance(real_G, LeapMotion_G)
distance_Grab_R_L_R = dtw.distance(real_R, LeapMotion_R)
distance_Grab_R_L_B = dtw.distance(real_B, LeapMotion_B)
distance_Grab_R_L_Y = dtw.distance(real_Y, LeapMotion_Y)

distance_Grab_R_T_G = dtw.distance(real_G, touch_G)
distance_Grab_R_T_R = dtw.distance(real_R, touch_R)
distance_Grab_R_T_B = dtw.distance(real_B, touch_B)
distance_Grab_R_T_Y = dtw.distance(real_Y, touch_Y)

distance_Grab_R_G_G = dtw.distance(real_G, glove_G)
distance_Grab_R_G_R = dtw.distance(real_R, glove_R)
distance_Grab_R_G_B = dtw.distance(real_B, glove_B)
distance_Grab_R_G_Y = dtw.distance(real_Y, glove_Y)

distance_Grab_R_H_G = dtw.distance(real_G, haptic_G)
distance_Grab_R_H_R = dtw.distance(real_R, haptic_R)
distance_Grab_R_H_B = dtw.distance(real_B, haptic_B)
distance_Grab_R_H_Y = dtw.distance(real_Y, haptic_Y)

print("[초]실졔vs립모션: " + str(distance_Grab_R_L_G))
print("[초]실졔vs 터치: " + str(distance_Grab_R_T_G))
print("[초]실졔vs 장갑: " + str(distance_Grab_R_G_G))
print("[초]실졔vs 햅틱: " + str(distance_Grab_R_H_G))
print()
print("[빨]실졔vs립모션: " + str(distance_Grab_R_L_R))
print("[빨]실졔vs 터치: " + str(distance_Grab_R_T_R))
print("[빨]실졔vs 장갑: " + str(distance_Grab_R_G_R))
print("[빨]실졔vs 햅틱: " + str(distance_Grab_R_H_R))
print()
print("[파]실졔vs립모션: " + str(distance_Grab_R_L_B))
print("[파]실졔vs 터치: " + str(distance_Grab_R_T_B))
print("[파]실졔vs 장갑: " + str(distance_Grab_R_G_B))
print("[파]실졔vs 햅틱: " + str(distance_Grab_R_H_B))
print()
print("[노]실졔vs립모션: " + str(distance_Grab_R_L_Y))
print("[노]실졔vs 터치: " + str(distance_Grab_R_T_Y))
print("[노]실졔vs 장갑: " + str(distance_Grab_R_G_Y))
print("[노]실졔vs 햅틱: " + str(distance_Grab_R_H_Y))
from dtaidistance import dtw_visualisation as dtwvis
'''
dtwvis.plot_warping(real_G, LeapMotion_G, path = dtw.warping_path(real_G, LeapMotion_G))
dtwvis.plot_warping(real_G, touch_G, path = dtw.warping_path(real_G, touch_G))
dtwvis.plot_warping(real_G, glove_G, path = dtw.warping_path(real_G, glove_G))
dtwvis.plot_warping(real_G, haptic_G, path = dtw.warping_path(real_G, haptic_G))
'''
'''

dtwvis.plot_warping(real_R, LeapMotion_R, path = dtw.warping_path(real_R, LeapMotion_R))
dtwvis.plot_warping(real_R, touch_R, path = dtw.warping_path(real_R, touch_R))
dtwvis.plot_warping(real_R, glove_R, path = dtw.warping_path(real_R, glove_R))
dtwvis.plot_warping(real_R, haptic_R, path = dtw.warping_path(real_R, haptic_R))
'''

'''
dtwvis.plot_warping(real_B, LeapMotion_B, path = dtw.warping_path(real_B, LeapMotion_B))
dtwvis.plot_warping(real_B, touch_B, path = dtw.warping_path(real_B, touch_B))
dtwvis.plot_warping(real_B, glove_B, path = dtw.warping_path(real_B, glove_B))
dtwvis.plot_warping(real_B, haptic_B, path = dtw.warping_path(real_B, haptic_B))
'''

dtwvis.plot_warping(real_Y, LeapMotion_Y, path = dtw.warping_path(real_Y, LeapMotion_Y))
dtwvis.plot_warping(real_Y, touch_Y, path = dtw.warping_path(real_Y, touch_Y))
dtwvis.plot_warping(real_Y, glove_Y, path = dtw.warping_path(real_Y, glove_Y))
dtwvis.plot_warping(real_Y, haptic_Y, path = dtw.warping_path(real_Y, haptic_Y))
''''''
#plt.plot([1, 2, 3, 4])
plt.show()
